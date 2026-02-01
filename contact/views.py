from django.shortcuts import render, redirect
from django.conf import settings
from django.contrib import messages

from openpyxl import Workbook, load_workbook
import os

def contact_page(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        subject = request.POST.get('subject')
        message = request.POST.get('message')

        # Excel file path
        file_path = os.path.join(settings.BASE_DIR, 'contact_messages.xlsx')

        # File exists असेल तर open, नाहीतर create
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Name", "Email", "Subject", "Message"])

        # Data add
        ws.append([name, email, subject, message])
        wb.save(file_path)

        messages.success(request, "Message sent successfully!")
        return redirect('contact')

    return render(request, 'contact/contact.html')